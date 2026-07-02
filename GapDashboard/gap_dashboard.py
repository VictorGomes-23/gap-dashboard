"""
gap_dashboard.py — Core pipeline for the County Gap Tracker

This module does NOT generate the HTML dashboard. The dashboard (index.html)
is a static file that lives in your git repo and rarely changes. This module's
only job is to read the live spreadsheet and write a small data.json file
that the dashboard fetches at runtime.

Can be run directly for a one-time manual build:ssessed
    python gap_dashboard.py              (build + push)
    python gap_dashboard.py --no-push     (build only, skip git push)
    python gap_dashboard.py --rebuild-history   (re-scan all of past/ to rebuild gap_history.json)

Normally you do NOT run this directly — watcher.py calls into it automatically
whenever the live spreadsheet changes.
"""

import os, sys, shutil, json, subprocess, time, argparse
import re
from datetime import datetime
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit(
        "openpyxl is not installed.\n"
        "Run:  pip install -r requirements.txt\n"
        "(from this script's folder) and try again."
    )

# =============================================================================
# CONFIGURATION — edit these four paths for your machine
# =============================================================================

# The live spreadsheet that analysts are editing right now.
LIVE_SOURCE = Path(r"C:\Users\victor.gomes\Fidelity National Financial\Grubowski, Alexa - test\Master-Gap Breakdown.xlsm")

# Your existing OneDrive working folder — current/, past/, and history live here.
WORK_DIR  = Path(r"C:\Documents\GapDashboardSite\GapDashboard")
DIR_CUR   = WORK_DIR / "current"
DIR_PAST  = WORK_DIR / "past"
HIST_FILE = WORK_DIR / "gap_history.json"

# The separate git repo folder that gets published to GitHub Pages.
# This must NOT be inside OneDrive — keep git and OneDrive sync apart.
REPO_DIR  = Path(r"C:\Documents\GapDashboardSite")
DATA_FILE = REPO_DIR / "data.json"

# =============================================================================
# HEADER-NAME-BASED COLUMN MAPPING
# Columns are matched by their header TEXT, not position. If the spreadsheet
# gains, loses, or reorders columns again in the future, this keeps working
# without any code changes.
# =============================================================================
HEADER_ALIASES = {
    "escalated": ["escalated", "escalation", "escalated?"],
    "code":      ["county code"],
    "state":     ["state"],
    "county":    ["county"],
    "gaps":      ["total gaps"],
    "volume":    ["volume level", "volume"],
    "gap_type":  ["gap types identified", "gap types"],
    "status":    ["current status", "status"],
    "action":    ["action required"],
    "comments":  ["notable comments / risks", "notable comments/risks", "comments", "notes"],
    "analyst":   ["analyst"],
    "assessed":  ["date last assessed by analyst", "date last assessed", "last assessed"],
}
def find_header_row_and_map(rows):
    """Scan the first several rows for a header row, matching by column TEXT.
    Returns (header_row_index, {field_name: column_index})."""


    def normalize_header(v):
        s = str(v or "")
        s = s.replace("\n", " ").replace("\r", " ")
        s = s.replace("\xa0", " ")  # handles Excel weird spaces
        s = re.sub(r"\s+", " ", s)  # ✅ collapse ALL multiple spaces into ONE
        return s.strip().lower()


    for i, row in enumerate(rows[:9]):
        cells = [normalize_header(c) for c in row]
        col_map, matches = {}, 0

        # ✅ PASS 1 — EXACT MATCHES ONLY (fixes your bug)
        for field, aliases in HEADER_ALIASES.items():
            for idx, cell in enumerate(cells):
                for alias in aliases:
                    if cell == alias:
                        col_map[field] = idx
                        matches += 1
                        break
                if field in col_map:
                    break

        # ✅ PASS 2 — PARTIAL MATCH (ONLY if not already found)
        for field, aliases in HEADER_ALIASES.items():
            if field in col_map:
                continue

            for idx, cell in enumerate(cells):
                for alias in aliases:
                    if alias in cell:
                        col_map[field] = idx
                        matches += 1
                        break
                if field in col_map:
                    break

        # ✅ Detect header row
        if matches >= 5:
            print("✅ Column Map:", col_map)  # optional debug (leave for now)
            return i, col_map

    return 0, {}
# =============================================================================
# UTILITIES
# =============================================================================
def to_date_str(val):
    if val is None:
        return None

    from datetime import datetime, timedelta

    # ✅ Case 1: datetime object
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    # ✅ Case 2: Excel numeric date
    try:
        f = float(val)
        if 40000 < f < 60000:
            return (datetime(1899, 12, 30) + timedelta(days=f)).strftime("%Y-%m-%d")
    except:
        pass

    # ✅ Case 3: string cleanup (CRITICAL FIX)
    s = str(val)

    # remove weird whitespace + line breaks
    s = s.replace("\n", " ").replace("\r", " ").replace("\xa0", " ").strip()

    if not s:
        return None

    # remove time portion safely
    parts = s.split()
    if len(parts) >= 1:
        s = parts[0]

    # normalize separators
    s = s.replace("-", "/")

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            continue

    return None
def days_since(ds):
    if not ds:
        return 999
    try:
        days = (datetime.today() - datetime.strptime(ds,"%Y-%m-%d")).days

        # ✅ clamp bad values
        if days < 0:
            return 0   # future dates → treat as fresh

        if days > 365:
            return 365  # cap outliers

        return days
    except:
        return 999

def safe_int(v):
    try:
        s = str(v).replace(",","").strip()
        if s.lower() in ("","none","inactive","unknown","n/a"): return 0
        return int(float(s))
    except: return 0


def clean(v, d=""):
    s = str(v or "").replace("\n", " ").replace("\r", " ").strip()
    return d if s.lower() in ("none","","nan") else s

def parse_escalated(v):
    if isinstance(v, bool): return v
    s = str(v or "").strip().lower()
    return s in ("yes","y","true","1","x","escalated","✓","✔")

def risk_score(r):
    """0–100 composite score. Components sum to a max of 100:
    gaps (35) + days stale (25) + volume (15) + status (10) + escalated (15)."""
    g   = min(r["gaps"]/10000*35, 35)
    st  = min((r["days_stale"] if r["days_stale"]<999 else 90)/60*25, 25)
    vo  = {"High":15,"Medium":8,"Low":0}.get(r["volume"],0)
    sc  = {"Active gaps":10,"Monitoring":5}.get(r["status"],0)
    esc = 15 if r.get("escalated") else 0
    return round(min(g+st+vo+sc+esc, 100))

def setup_dirs():
    for d in [DIR_CUR, DIR_PAST]:
        d.mkdir(parents=True, exist_ok=True)
    REPO_DIR.mkdir(parents=True, exist_ok=True)

# =============================================================================
# FILE SYNC — copy-based, not move-based, since the source is live and
# being edited by other people. Archives to past/ once per calendar day
# (matched against the source file's own last-modified date), and simply
# overwrites current/ for any other updates that happen within the same day.
# This keeps past/ from filling up with near-duplicate snapshots while still
# preserving exactly one trend data point per day.
# =============================================================================
def wait_for_stable_file(path, checks=3, interval=2, max_wait=40):
    """Wait until the file's size stops changing — avoids reading a file
    mid-save."""
    last_size, stable, waited = -1, 0, 0
    while waited < max_wait:
        try:
            size = path.stat().st_size
        except FileNotFoundError:
            return False
        if size == last_size:
            stable += 1
            if stable >= checks:
                return True
        else:
            stable = 0
        last_size = size
        time.sleep(interval)
        waited += interval
    return False

def sync_current_from_source():
    if not LIVE_SOURCE.exists():
        print(f"[ERROR] Live source file not found:\n  {LIVE_SOURCE}")
        return False

    if not wait_for_stable_file(LIVE_SOURCE):
        print("[WARN] File still changing after max wait — proceeding anyway.")

    today_str = datetime.today().strftime("%Y%m%d")
    cur_files = list(DIR_CUR.glob("*.xls*"))

    if cur_files:
        existing = cur_files[0]
        existing_day = datetime.fromtimestamp(existing.stat().st_mtime).strftime("%Y%m%d")
        if existing_day != today_str:
            ts = datetime.fromtimestamp(existing.stat().st_mtime).strftime("%Y%m%d_%H%M%S")
            dest = DIR_PAST / f"{existing.stem}_{ts}{existing.suffix}"
            shutil.move(str(existing), str(dest))
            print(f"[OK] New day — archived previous snapshot -> past/{dest.name}")
        else:
            existing.unlink()

    dest_cur = DIR_CUR / LIVE_SOURCE.name
    for attempt in range(5):
        try:
            shutil.copy2(str(LIVE_SOURCE), str(dest_cur))
            print(f"[OK] Synced current/ <- {LIVE_SOURCE.name}")
            return True
        except PermissionError:
            print(f"[WARN] File locked, retrying ({attempt+1}/5)...")
            time.sleep(4)
    print("[ERROR] Could not copy source file after retries.")
    return False

# =============================================================================
# SPREADSHEET PARSING
# =============================================================================
def parse_xlsx(path):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if "cover" in s.lower()), wb.active)
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    hi, cmap = find_header_row_and_map(rows)
    if not cmap:
        raise ValueError("Could not locate a recognizable header row in the spreadsheet.")

    def col(row, field, default=""):
        idx = cmap.get(field, -1)
        return row[idx] if 0 <= idx < len(row) else default

    # ✅ GAP RECOVERY
    def extract_gaps(row, fallback):
        val = safe_int(fallback)
        if val > 0:
            return val

        candidates = []
        for c in row:
            try:
                n = int(str(c).replace(",", "").strip())
                if 0 < n < 100000000:
                    candidates.append(n)
            except:
                continue

        return max(candidates) if candidates else 0

    # ✅ STATUS NORMALIZATION
    def normalize_status(val, row):
        s = clean(val).lower()

        if "active" in s:
            return "Active gaps"
        if "monitor" in s:
            return "Monitoring"
        if "pending" in s:
            return "Pending"
        if "clean" in s:
            return "Clean"
        if "inactive" in s:
            return "Inactive"

        # fallback scan entire row
        for c in row:
            sc = str(c).lower()
            if "active" in sc:
                return "Active gaps"
            if "monitor" in sc:
                return "Monitoring"
            if "pending" in sc:
                return "Pending"
            if "clean" in sc:
                return "Clean"

        return "Unknown"

    records = []
    last_assessed = None
    for row in rows[hi+1:]:
        if not any(row):
            continue

        row = list(row)

        state = clean(col(row, "state"))
        county = clean(col(row, "county"))
        if not state or len(state) > 4 or state.upper() in ("STATE", "NONE"):
            continue

        raw_gaps = col(row, "gaps", 0)
        gaps = extract_gaps(row, raw_gaps)

        inactive = str(raw_gaps).strip().lower() == "inactive"

        raw_status = col(row, "status")
        status = "Inactive" if inactive else normalize_status(raw_status, row)

        analyst = clean(col(row, "analyst"), "Unassigned")
        if analyst.lower() in ("inactive", "none", "n/a", ""):
            analyst = "Unassigned"


        raw_assessed = col(row, "assessed")

# ✅ normalize
        raw_clean = str(raw_assessed).replace("\xa0", " ").strip()

        if raw_clean:
            assessed = to_date_str(raw_assessed)
        else:
    # ✅ HARD FALLBACK — scan entire row for ANY date
            assessed = None
            for c in row:
                a = to_date_str(c)
                if a:
                    assessed = a
                    break

# ✅ FINAL safety fallback (optional but recommended)
        if assessed is None:
            assessed = "1900-01-01"  # or leave None depending on preference

        stale = days_since(assessed)



        r = {
            "code": clean(col(row, "code")),
            "state": state,
            "county": county,
            "gaps": gaps,
            "volume": clean(col(row, "volume"), "Low"),
            "gap_type": clean(col(row, "gap_type")),
            "status": status,
            "action": clean(col(row, "action")),
            "comments": clean(col(row, "comments")),
            "analyst": analyst,
            "assessed": assessed,
            "days_stale": stale,
            "escalated": parse_escalated(col(row, "escalated")),
        }

        r["risk"] = risk_score(r)
        records.append(r)

    return records
def parse_xlsx_with_retry(path, attempts=5, delay=5):
    for i in range(attempts):
        try:
            return parse_xlsx(path)
        except Exception as e:
            print(f"[WARN] Parse attempt {i+1}/{attempts} failed: {e}")
            time.sleep(delay)
    print("[ERROR] Could not parse spreadsheet after multiple attempts.")
    return None


# =============================================================================
# HISTORY (trend snapshots — one entry per calendar date, refined throughout
# the day as the watcher runs). enrich_from_past() re-scans every archived
# file and is intentionally NOT called automatically every cycle — only via
# --rebuild-history — since past/ can grow large under continuous watching
# and re-parsing it every run would get slower over time for no benefit.
# =============================================================================
def load_history():
    if HIST_FILE.exists():
        try:
            with open(HIST_FILE, encoding="utf-8") as f: return json.load(f)
        except: pass
    return []

def save_history(h):
    with open(HIST_FILE, "w", encoding="utf-8") as f:
        json.dump(h, f, indent=2, default=str)

def make_snapshot(records, ds):
    ba = defaultdict(lambda: {"gaps":0,"active":0,"clean":0})
    for r in records:
        a = r["analyst"]; ba[a]["gaps"] += r["gaps"]
        if r["status"].lower().startswith("active"): ba[a]["active"] += 1
        if r["status"].lower().startswith("clean"):       ba[a]["clean"]  += 1
    return {
        "date": ds,
        "total_gaps": sum(r["gaps"] for r in records),
        "total_counties": len(records),
        "clean": sum(1 for r in records if r["status"].lower().startswith("clean")),
        "active": sum(1 for r in records if r["status"].lower().startswith("active")),
        "monitoring": sum(1 for r in records if r["status"].lower().startswith("monitoring")),
        "pending": sum(1 for r in records if r["status"].lower().startswith("pending")),
        "high_vol_active": sum(1 for r in records if r["status"].lower().startswith("active") and r["volume"]=="High"),
        "escalated": sum(1 for r in records if r.get("escalated")),
        "by_analyst": {k: dict(v) for k,v in ba.items()},
    }

def upsert_history(history, records, ds):
    snap = make_snapshot(records, ds)
    dates = [h["date"] for h in history]
    if ds in dates: history[dates.index(ds)] = snap
    else: history.append(snap)
    history.sort(key=lambda x: x["date"])
    return history

def enrich_from_past(history):
    """Manual recovery tool — re-parses every archived file in past/ to
    rebuild gap_history.json from scratch. Run via --rebuild-history."""
    existing = {h["date"] for h in history}
    for pf in sorted(list(DIR_PAST.glob("*.xls*")), key=lambda p: p.name):
        pf_date = None
        for part in pf.stem.split("_"):
            if len(part)==8 and part.isdigit():
                try: pf_date = datetime.strptime(part,"%Y%m%d").strftime("%Y-%m-%d"); break
                except: pass
        if not pf_date or pf_date in existing: continue
        try:
            pr = parse_xlsx(pf)
            history = upsert_history(history, pr, pf_date)
            existing.add(pf_date)
            print(f"[OK] History rebuilt: {pf.name} ({pf_date})")
        except Exception as e:
            print(f"[WARN] {pf.name}: {e}")
    return history

# =============================================================================
# COMPUTE DASHBOARD DATA
# =============================================================================
def compute_data(records):
    # NOTE: this only computes what the dashboard's own JS can't cheaply derive
    # itself from `records` on every filter change. Earlier versions also built
    # critical_counties/on_hold_counties/stale_counties/escalated_counties/
    # top_counties/analyst_queues here and shipped them in data.json, but
    # index.html never reads any of those — it recomputes the same lists
    # client-side from `records` on every filter interaction anyway. Keeping
    # both meant every poll shipped duplicate copies of up to ~65 full county
    # records for nothing. Only the counts survive here, for the console log.
    total = len(records); tg = sum(r["gaps"] for r in records)
    sc = defaultdict(int)
    for r in records: sc[r["status"]] += 1
    health = round(sc.get("Clean",0)/total*100,1) if total else 0
    escalated_count = sum(1 for r in records if r.get("escalated"))

    bs = defaultdict(lambda: {"gaps":0,"active":0,"clean":0,"monitoring":0,"counties":0})
    for r in records:
        s = r["state"]; bs[s]["gaps"] += r["gaps"]; bs[s]["counties"] += 1
        if r["status"].lower().startswith("active"): bs[s]["active"] += 1
        elif r["status"].lower().startswith("clean"):     bs[s]["clean"]  += 1
        elif r["status"].lower().startswith("monitoring"): bs[s]["monitoring"] += 1
    state_list = sorted([{"state":k,**dict(v)} for k,v in bs.items()], key=lambda x:-x["gaps"])

    analysts = sorted({r["analyst"] for r in records if r["analyst"] not in ("Unassigned","")})
    ba = defaultdict(lambda: {"total":0,"gaps":0,"active":0,"clean":0,"monitoring":0,"pending":0})
    for r in records:
        a = r["analyst"]; ba[a]["total"] += 1; ba[a]["gaps"] += r["gaps"]
        ba[a]["active"]     += 1 if r["status"].lower().startswith("active") else 0
        ba[a]["clean"]      += 1 if r["status"].lower().startswith("clean")       else 0
        ba[a]["monitoring"] += 1 if r["status"].lower().startswith("monitoring")  else 0
        ba[a]["pending"]    += 1 if r["status"].lower().startswith("pending")     else 0

    gtc = defaultdict(int)
    for r in records:
        gt = r["gap_type"].strip()
        if gt and gt.lower() not in ("","none"," "): gtc[gt] += 1

    return {
        "total_counties": total, "total_gaps": tg, "health_score": health,
        "status_counts": dict(sc), "escalated_count": escalated_count,
        "state_list": state_list, "analysts": analysts,
        "by_analyst": {k:dict(v) for k,v in ba.items()},
        "gap_types": dict(gtc), "all_records": records,
    }

def build_payload(data, history, source_name):
    return {
        "meta": {
            "file": source_name,
            "date": datetime.today().strftime("%Y-%m-%d"),
            "generated": datetime.today().strftime("%Y-%m-%d %H:%M:%S"),
        },
        "kpis": {k:v for k,v in data.items() if k != "all_records"},
        "history": history,
        "records": data["all_records"],
    }

# =============================================================================
# GIT PUSH
# =============================================================================
_NO_WINDOW = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0

def _git(*args):
    return subprocess.run(["git", *args], cwd=REPO_DIR, check=True,
                           capture_output=True, creationflags=_NO_WINDOW)

def git_push(message):
    if not (REPO_DIR / ".git").exists():
        print(f"[ERROR] {REPO_DIR} is not a git repo. See setup instructions.")
        return False
    try:
        _git("add", "-A")
        status = subprocess.run(["git","status","--porcelain"], cwd=REPO_DIR, capture_output=True, text=True, creationflags=_NO_WINDOW)
        if not status.stdout.strip():
            print("[INFO] No changes to publish.")
            return False
        _git("commit", "-m", message)

        # Rebase onto whatever's on the remote before pushing. Without this,
        # any commit made to the repo from somewhere else (another machine, a
        # PR merge, a manual edit) makes this push a non-fast-forward reject,
        # which previously required a manual `git pull` + merge to unstick —
        # this is almost certainly why past history shows a "Complete merge"
        # commit in the middle of otherwise-automatic updates.
        try:
            _git("pull", "--rebase", "--autostash")
        except subprocess.CalledProcessError as e:
            err = e.stderr.decode() if e.stderr else str(e)
            print(f"[ERROR] git pull --rebase failed, not pushing: {err}")
            return False

        _git("push")
        print("[OK] Pushed to GitHub Pages.")
        return True
    except subprocess.CalledProcessError as e:
        err = e.stderr.decode() if e.stderr else str(e)
        print(f"[ERROR] git push failed: {err}")
        return False

# =============================================================================
# PIPELINE ORCHESTRATION
# =============================================================================
def run_pipeline(push=True):
    setup_dirs()
    if not sync_current_from_source():
        return False

    cur_files = list(DIR_CUR.glob("*.xls*"))
    if not cur_files:
        print("[ERROR] No spreadsheet in current/ after sync."); return False
    cur_file = cur_files[0]

    print(f"> Parsing {cur_file.name}...")
    records = parse_xlsx_with_retry(cur_file)
    if records is None:
        return False
    print(f"  {len(records)} county records loaded")

    history = load_history()
    history = upsert_history(history, records, datetime.today().strftime("%Y-%m-%d"))
    save_history(history)

    data = compute_data(records)
    print(f"  Health: {data['health_score']}%  Gaps: {data['total_gaps']:,}  "
          f"Active: {data['status_counts'].get('Active gaps',0)}  "
          f"Escalated: {data['escalated_count']}")

    payload = build_payload(data, history, LIVE_SOURCE.name)
    REPO_DIR.mkdir(parents=True, exist_ok=True)
    DATA_FILE.write_text(json.dumps(payload, default=str, ensure_ascii=False), encoding="utf-8")
    print(f"[OK] data.json written ({DATA_FILE.stat().st_size:,} bytes)")

    if push:
        git_push(f"Update gap data — {datetime.today().strftime('%Y-%m-%d %H:%M')}")
    return True

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-push", action="store_true", help="Build data.json but skip git push")
    ap.add_argument("--rebuild-history", action="store_true", help="Re-scan past/ to rebuild gap_history.json")
    args = ap.parse_args()

    if args.rebuild_history:
        setup_dirs()
        history = load_history()
        history = enrich_from_past(history)
        save_history(history)
        print(f"[OK] History rebuilt — {len(history)} snapshot(s) total.")
        return

    run_pipeline(push=not args.no_push)

if __name__ == "__main__":
    main()